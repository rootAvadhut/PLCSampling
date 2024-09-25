import asyncio
from pymodbus.client import AsyncModbusTcpClient
from pymodbus.exceptions import ConnectionException
import socket
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import home_ui # Import from your ui.py module
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

# Global flag to track the connection status
connected = False

# Path to the Excel file
EXCEL_FILE_PATH = r"E:\project 4\24-09-2024\temp\analog_register_data.xlsx"

# Global variable to store the previous register values
previous_input_registers = None

# Ensure directory exists before saving the file
def ensure_directory_exists(file_path):
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)

# Function to check if the Excel file is open
def is_excel_file_open(file_path):
    try:
        # Try to open the file in append mode
        with open(file_path, 'a'):
            return False  # File is not open
    except IOError:
        return True  # File is open
def write_to_excel(input_registers, ip_address, port_number):
    if len(input_registers) < 8:
        input_registers += [None] * (8 - len(input_registers))

    current_time = datetime.now()
    date = current_time.strftime("%Y-%m-%d")
    time = current_time.strftime("%H:%M")

    data = [date, time,ip_address, port_number] + input_registers[:8]

    print(f"EXCEL_FILE_PATH: {EXCEL_FILE_PATH}")
    
    if is_excel_file_open(EXCEL_FILE_PATH):
        print("Excel file is currently open. Please close it to write data.")
        return

    if os.path.exists(EXCEL_FILE_PATH):
        try:
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
        except (InvalidFileException, BadZipFile):
            print(f"File {EXCEL_FILE_PATH} is corrupted or not a valid Excel file. Deleting and recreating.")
            os.remove(EXCEL_FILE_PATH)
            wb = Workbook()
            ws = wb.active
            ws.title = "Analog Data"
            ws.append([
                "Date", "Time","IP Address", "Port Number", "Register 300002", "Register 300003", "Register 300004",
                "Register 300005", "Register 300006", "Register 300007",
                "Register 300008", "Register 300009"
            ])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Analog Data"
        ws.append([
            "Date", "Time","IP Address", "Port Number", "Register 300002", "Register 300003", "Register 300004",
            "Register 300005", "Register 300006", "Register 300007",
            "Register 300008", "Register 300009"
        ])

    try:
        ws.append(data)
        wb.save(EXCEL_FILE_PATH)
        print("Data written to Excel file.")
    except PermissionError:
        print("Permission denied: The Excel file is currently opened. Please close it to write data.")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")

# # Updated write_to_excel function to accept three arguments
# def write_to_excel(input_registers, ip_address, port_number):
#     if len(input_registers) < 8:
#         input_registers += [None] * (8 - len(input_registers))

#     current_time = datetime.now()
#     date = current_time.strftime("%Y-%m-%d")
#     time = current_time.strftime("%H:%M")

#     # Include IP address and port number in the data to be written
#     data = [date, time, ip_address, port_number] + input_registers[:8]

#     print(f"EXCEL_FILE_PATH: {EXCEL_FILE_PATH}")
#     if os.path.exists(EXCEL_FILE_PATH):
#         try:
#             wb = load_workbook(EXCEL_FILE_PATH)
#         except (InvalidFileException, BadZipFile):
#             print(f"File {EXCEL_FILE_PATH} is corrupted or not a valid Excel file. Deleting and recreating.")
#             os.remove(EXCEL_FILE_PATH)
#             wb = Workbook()
#             ws = wb.active
#             ws.title = "Analog Data"
#             ws.append([
#                 "Date", "Time", "IP Address", "Port Number",
#                 "Register 300002", "Register 300003", "Register 300004",
#                 "Register 300005", "Register 300006", "Register 300007",
#                 "Register 300008", "Register 300009"
#             ])
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Analog Data"
#         ws.append([
#             "Date", "Time", "IP Address", "Port Number",
#             "Register 300002", "Register 300003", "Register 300004",
#             "Register 300005", "Register 300006", "Register 300007",
#             "Register 300008", "Register 300009"
#         ])

#     ws = wb.active
#     ws.append(data)
#     wb.save(EXCEL_FILE_PATH)
#     print("Data written to Excel file.")

# Function to check if data has changed
def has_data_changed(current_input_registers):
    global previous_input_registers
    if previous_input_registers is None:
        previous_input_registers = current_input_registers
        return True

    if current_input_registers != previous_input_registers:
        previous_input_registers = current_input_registers
        return True

    return False

# Function to check the connection status
async def check_connection(client):
    global connected
    try:
        coil_check = await client.read_coils(0, 1)
        coil_ok = not coil_check.isError()

        input_register_check = await client.read_input_registers(0, 1)
        input_register_ok = not input_register_check.isError()

        input_status_check = await client.read_discrete_inputs(0, 1)
        input_status_ok = not input_status_check.isError()

        if coil_ok or input_register_ok or input_status_ok:
            connected = True
            return True
        else:
            connected = False
            return False

    except (ConnectionException, socket.error) as e:
        print(f"Connection error: {e}")
        connected = False
        return False
    except Exception as e:
        print(f"An unexpected error occurred during the connection check: {e}")
        connected = False
        return False

# Function to read coil states
async def run_modbus_client_coil(client, sampling_frequency):
    try:
        digital_result = await client.read_coils(0, 15)
        if digital_result.isError():
            return None
        else:
            await asyncio.sleep(sampling_frequency / 1000)
            return digital_result.bits
    except Exception as e:
        print(f"Error reading coils: {e}")
        return None

# Function to read input registers
async def run_modbus_client_input_register(client, sampling_frequency, ip_address, port_number):
    try:
        input_register_result = await client.read_input_registers(0, 8)
        if input_register_result.isError():
            return None
        else:
            current_input_registers = input_register_result.registers

            if has_data_changed(current_input_registers):
                asyncio.create_task(async_write_to_excel(current_input_registers, ip_address, port_number))

            await asyncio.sleep(sampling_frequency / 1000)
            return current_input_registers
    except Exception as e:
        print(f"Error reading input registers: {e}")
        return None

# Function to write to Excel asynchronously
async def async_write_to_excel(input_registers, ip_address, port_number):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, write_to_excel, input_registers, ip_address, port_number)

# Function to read input statuses
async def run_modbus_client_input_status(client, sampling_frequency):
    try:
        input_status_result = await client.read_discrete_inputs(0, 15)
        if input_status_result.isError():
            return None
        else:
            await asyncio.sleep(sampling_frequency / 1000)
            return input_status_result.bits
    except Exception as e:
        print(f"Error reading input statuses: {e}")
        return None

# Main Modbus client loop
async def modbus_client_loop(ip_address, port, sampling_frequency, update_ui_callback):
    global connected
    async with AsyncModbusTcpClient(ip_address, port=port) as client:
        while True:
            if not connected:
                if await check_connection(client):
                    print("Machine connected. Starting data retrieval...")
                else:
                    await asyncio.sleep(sampling_frequency / 1000)
                    continue

            coil_states = await run_modbus_client_coil(client, sampling_frequency)
            input_registers = await run_modbus_client_input_register(client, sampling_frequency, ip_address, port)
            input_statuses = await run_modbus_client_input_status(client, sampling_frequency)

            update_ui_callback(coil_states, input_registers, input_statuses)

# Function to start the Modbus client
def run_client(ip_address, port, sampling_frequency, update_ui_callback):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(modbus_client_loop(ip_address, port, sampling_frequency, update_ui_callback))
    loop.close()








# import asyncio
# from pymodbus.client import AsyncModbusTcpClient
# from pymodbus.exceptions import ConnectionException
# import socket
# import os
# from openpyxl import Workbook, load_workbook
# from datetime import datetime

# # Global flag to track the connection status
# connected = False

# # Path to the Excel file
# EXCEL_FILE_PATH = "analog_register_data.xlsx"

# # Global variable to store the previous register values
# previous_input_registers = None

# # Ensure directory exists before saving the file
# def ensure_directory_exists(file_path):
#     directory = os.path.dirname(file_path)
#     if directory and not os.path.exists(directory):
#         os.makedirs(directory)

# # Function to write data into an Excel file
# def write_to_excel(input_registers):
#     if len(input_registers) < 8:
#         input_registers += [None] * (8 - len(input_registers))

#     current_time = datetime.now()
#     date = current_time.strftime("%Y-%m-%d")
#     time = current_time.strftime("%H:%M")

#     data = [date, time] + input_registers[:8]

#     ensure_directory_exists(EXCEL_FILE_PATH)

#     if not os.path.exists(EXCEL_FILE_PATH):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Analog Data"
#         ws.append([
#             "Date", "Time", "Register 300002", "Register 300003", "Register 300004",
#             "Register 300005", "Register 300006", "Register 300007",
#             "Register 300008", "Register 300009"
#         ])
#     else:
#         wb = load_workbook(EXCEL_FILE_PATH)
#         ws = wb.active

#     ws.append(data)
#     wb.save(EXCEL_FILE_PATH)
#     print("Data written to Excel file.")

# # Function to check if data has changed
# def has_data_changed(current_input_registers):
#     global previous_input_registers
#     if previous_input_registers is None:
#         previous_input_registers = current_input_registers
#         return True

#     if current_input_registers != previous_input_registers:
#         previous_input_registers = current_input_registers
#         return True

#     return False

# # Function to check the connection status
# async def check_connection(client):
#     global connected
#     try:
#         coil_check = await client.read_coils(0, 1)
#         coil_ok = not coil_check.isError()

#         input_register_check = await client.read_input_registers(0, 1)
#         input_register_ok = not input_register_check.isError()

#         input_status_check = await client.read_discrete_inputs(0, 1)
#         input_status_ok = not input_status_check.isError()

#         if coil_ok or input_register_ok or input_status_ok:
#             connected = True
#             return True
#         else:
#             connected = False
#             return False

#     except (ConnectionException, socket.error) as e:
#         print(f"Connection error: {e}")
#         connected = False
#         return False
#     except Exception as e:
#         print(f"An unexpected error occurred during the connection check: {e}")
#         connected = False
#         return False

# # Function to read coil states
# async def run_modbus_client_coil(client, sampling_frequency):
#     try:
#         digital_result = await client.read_coils(0, 15)
#         if digital_result.isError():
#             return None
#         else:
#             await asyncio.sleep(sampling_frequency / 1000)
#             return digital_result.bits
#     except Exception as e:
#         print(f"Error reading coils: {e}")
#         return None

# # Function to read input registers
# async def run_modbus_client_input_register(client, sampling_frequency):
#     try:
#         input_register_result = await client.read_input_registers(0, 8)
#         if input_register_result.isError():
#             return None
#         else:
#             current_input_registers = input_register_result.registers

#             if has_data_changed(current_input_registers):
#                 asyncio.create_task(async_write_to_excel(current_input_registers))

#             await asyncio.sleep(sampling_frequency / 1000)
#             return current_input_registers
#     except Exception as e:
#         print(f"Error reading input registers: {e}")
#         return None

# # Function to write to Excel asynchronously
# async def async_write_to_excel(input_registers):
#     loop = asyncio.get_event_loop()
#     await loop.run_in_executor(None, write_to_excel, input_registers)

# # Function to read input statuses
# async def run_modbus_client_input_status(client, sampling_frequency):
#     try:
#         input_status_result = await client.read_discrete_inputs(0, 15)
#         if input_status_result.isError():
#             return None
#         else:
#             await asyncio.sleep(sampling_frequency / 1000)
#             return input_status_result.bits
#     except Exception as e:
#         print(f"Error reading input statuses: {e}")
#         return None

# # Main Modbus client loop
# async def modbus_client_loop(ip_address, port, sampling_frequency, update_ui_callback):
#     global connected
#     async with AsyncModbusTcpClient(ip_address, port=port) as client:
#         while True:
#             if not connected:
#                 if await check_connection(client):
#                     print("Machine connected. Starting data retrieval...")
#                 else:
#                     await asyncio.sleep(sampling_frequency / 1000)
#                     continue

#             coil_states = await run_modbus_client_coil(client, sampling_frequency)
#             input_registers = await run_modbus_client_input_register(client, sampling_frequency)
#             input_statuses = await run_modbus_client_input_status(client, sampling_frequency)

#             update_ui_callback(coil_states, input_registers, input_statuses)

#             # await asyncio.sleep(sampling_frequency / 1000)

# # Function to start the Modbus client
# def run_client(ip_address, port, sampling_frequency, update_ui_callback):
#     loop = asyncio.new_event_loop()
#     asyncio.set_event_loop(loop)
#     loop.run_until_complete(modbus_client_loop(ip_address, port, sampling_frequency, update_ui_callback))
#     loop.close()
